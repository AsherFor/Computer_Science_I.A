from tkinter import *
from matplotlib import pyplot as plt
import pandas as pd
import runpy
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import scrolledtext


main_array = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
column_titles_sheet1 = ['Name', 'Date', 'Notes', 'Driver', '3-Wood', '5-Wood', '3-Iron', '4-Iron', '5-Iron', '6-Iron', '7-Iron', '8-Iron', '9-Iron', 'Pitching Wedge', 'Gap Wedge', 'Sand Wedge', 'Lob Wedge', 'Putter']

#Loading my excel file if it is already created
try:
    wb = load_workbook('Golf_Statistics.xlsx')
    ws = wb.worksheets[0]

#Creates excel file
except FileNotFoundError:
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Average Distance of Golf Club"

    sheet1['A1'] = column_titles_sheet1[0]
    sheet1['B1'] = column_titles_sheet1[1]
    sheet1['C1'] = column_titles_sheet1[2]
    sheet1['D1'] = column_titles_sheet1[3]
    sheet1['E1'] = column_titles_sheet1[4]
    sheet1['F1'] = column_titles_sheet1[5]
    sheet1['G1'] = column_titles_sheet1[6]
    sheet1['H1'] = column_titles_sheet1[7]
    sheet1['I1'] = column_titles_sheet1[8]
    sheet1['J1'] = column_titles_sheet1[9]
    sheet1['K1'] = column_titles_sheet1[10]
    sheet1['L1'] = column_titles_sheet1[11]
    sheet1['M1'] = column_titles_sheet1[12]
    sheet1['N1'] = column_titles_sheet1[13]
    sheet1['O1'] = column_titles_sheet1[14]
    sheet1['P1'] = column_titles_sheet1[15]
    sheet1['Q1'] = column_titles_sheet1[16]
    sheet1['R1'] = column_titles_sheet1[17]

    wb.save(filename='Golf_Statistics.xlsx')

#Main page of my tkinter GUI
master = Tk()
master.title("Golf Statistics")
master.geometry("700x300")
master.configure(background="white")

#Header of the Program
main_header = Label(master, text="Golf Statistics")
main_header.config(font =("Times New Roman", 30))
main_header.grid(row=0, column=0)

#Entry for name
Label(master, text='First Name and Last Name').grid(row=2)
name = Entry(master)
name.grid(row=2, column=1)


def name_user():
    entry_name = name.get()
    main_array[0] = entry_name

#Entry for Date
Label(master, text='Date').grid(row=3)
date = Entry(master)
date.grid(row=3, column=1)

def date_of_user():
    entry_date = date.get()
    main_array[1] = entry_date

#Individual notes about golfer
Label(master, text='Individual notes about golfer').grid(row=8)
notes = Entry(master)
notes.grid(row=8, column=1)

def user_notes():
    entry_notes = notes.get()
    main_array[2] = entry_notes

#Opens new GUI window to calculate the average distance of a golf club
def average_distance():
    #Main window
    newWindow = Toplevel(master)
    newWindow.title("Calculate Average Distance")
    newWindow.geometry("900x200")

    #Header
    main_header = Label(newWindow, text="Average Distance of a Golf Club")
    main_header.config(font=("Times New Roman", 20))
    main_header.grid(row=0, column=0)

    #Buttons
    Button(newWindow, text="Graph", command= lambda: average_distance_bar_chart(distance, club_name)).grid(row=3, column=0)
    Button(newWindow, text="Calculate", command= lambda: claculate_average_distance(distance, club_name)).grid(row=3, column=1)

    # Label and entry for club and distance of golf shots
    Label(newWindow, text='Enter one of the following clubs (Driver, 3-Wood, 5-Wood, 3-Iron, 4-Iron, 5-Iron,\n'
                          '6-Iron, 7-Iron, 8-Iron, 9-Iron, Pitching Wedge, Gap Wdge, Sand Wedge, Lob Wedge, Putter).').grid(row=1)
    Label(newWindow, text='Enter Distance of Golf Shots (Ex: 400 + 200 + 100)').grid(row=2)
    club_name = Entry(newWindow)
    distance = Entry(newWindow)
    club_name.grid(row=1, column=1)
    distance.grid(row=2, column=1)

#Calculations of the average distance of a golf club
def claculate_average_distance(distance, club_name):
    entry_club_name = club_name.get()
    entry_distance = distance.get()

    # This is a for loop to count the amount of times the + character is in the entry for distance
    value = 0
    for x in entry_distance:
        if x == '+':
            value = value + 1
    number_of_symbols = value + 1
    average_yards = (eval(entry_distance) / number_of_symbols)
    print("You are able to hit your", entry_club_name, "an average distance of", round(average_yards, 2), "yards!")

   #The following if statements checks the club name the user entered matches a column title, and then assigns the average distance to the main array
    if entry_club_name == column_titles_sheet1[3]:
        main_array[3] = average_yards
    if entry_club_name == column_titles_sheet1[4]:
        main_array[4] = average_yards
    if entry_club_name == column_titles_sheet1[5]:
        main_array[5] = average_yards
    if entry_club_name == column_titles_sheet1[6]:
        main_array[6] = average_yards
    if entry_club_name == column_titles_sheet1[7]:
        main_array[7] = average_yards
    if entry_club_name == column_titles_sheet1[8]:
        main_array[8] = average_yards
    if entry_club_name == column_titles_sheet1[9]:
        main_array[9] = average_yards
    if entry_club_name == column_titles_sheet1[10]:
        main_array[10] = average_yards
    if entry_club_name == column_titles_sheet1[11]:
        main_array[11] = average_yards
    if entry_club_name == column_titles_sheet1[12]:
        main_array[12] = average_yards
    if entry_club_name == column_titles_sheet1[13]:
        main_array[13] = average_yards
    if entry_club_name == column_titles_sheet1[14]:
        main_array[14] = average_yards
    if entry_club_name == column_titles_sheet1[15]:
        main_array[15] = average_yards
    if entry_club_name == column_titles_sheet1[16]:
        main_array[16] = average_yards
    if entry_club_name == column_titles_sheet1[17]:
        main_array[17] = average_yards

#The bar chart for the average distance of a golf club
def average_distance_bar_chart(distance, club_name):
    entry_distance = distance.get()
    entry_club_name = club_name.get()

    # This is a for loop to count the amount of times the + character is in the entry for distance
    value = 0
    for x in entry_distance:
        if x == '+':
            value = value + 1
    number_of_symbols = value + 1
    average_yards = (eval(entry_distance) / number_of_symbols)

    plt.barh([entry_club_name], [round(average_yards, 2)], align='center', label="Distance")
    plt.legend()

    plt.xlim([0, 500])
    plt.ylabel('Type of Golf Club')
    plt.xlabel('Distance (yds)')
    plt.title('Average Distance of ' + entry_club_name)

    plt.show()

#Opens new GUI window to calculate the percentage chance of hitting a certain accuracy with a specific club
def accuracy_of_shot():
    #Main Window
    newWindow = Toplevel(master)
    newWindow.title("Calculate Accuracy")
    newWindow.geometry("700x200")

    #Header
    main_header = Label(newWindow, text="Accuracy of a Golf Club")
    main_header.config(font=("Times New Roman", 20))
    main_header.grid(row=0, column=0)

    #Buttons
    Button(newWindow, text="Graph", command= lambda: accuruacy_pie_chart(accuracy, name_of_club)).grid(row=3, column=0)
    Button(newWindow, text="Calculate", command= lambda: calculate_accuracy(name_of_club, accuracy)).grid(row=3, column=1)

    #Label and entry for accuracy of a shot
    Label(newWindow, text='What Club Are You Using?').grid(row=1)
    Label(newWindow, text='Enter the type of shot you hit (Hook, Slice, Fade, Draw, Push, Pull)').grid(row=2)
    name_of_club = Entry(newWindow)
    accuracy = Entry(newWindow)
    name_of_club.grid(row=1, column=1)
    accuracy.grid(row=2, column=1)

#Calculations of the accuracy of a golf club
def calculate_accuracy(name_of_club, accuracy):
    entry_name_of_club = name_of_club.get()
    entry_accuracy = accuracy.get()

    #Counting the amount of occurrences of each type of shot
    Hook = entry_accuracy.count("Hook")
    Slice = entry_accuracy.count("Slice")
    Fade = entry_accuracy.count("Fade")
    Draw = entry_accuracy.count("Draw")
    Push = entry_accuracy.count("Push")
    Pull = entry_accuracy.count("Pull")

    # Calculating the percentage of hitting a certain accuracy shot
    result_hook = (Hook / len(entry_accuracy.split()) * 100)
    result_slice = (Slice / len(entry_accuracy.split()) * 100)
    result_fade = (Fade / len(entry_accuracy.split()) * 100)
    result_draw = (Draw / len(entry_accuracy.split()) * 100)
    result_push = (Push / len(entry_accuracy.split()) * 100)
    result_pull = (Pull / len(entry_accuracy.split()) * 100)

    #If statements to check if a person inputed one of the type of accuracies, to then print the result of the calculation
    if Hook > 0:
        print("You will hit a hook shot with your", entry_name_of_club, round(result_hook, 2), "% of the time out of", len(entry_accuracy.split()), "shots!")
    if Slice > 0:
        print("You will hit a slice shot with your", entry_name_of_club, round(result_slice, 2), "%% of the time out of", len(entry_accuracy.split()), "shots!")
    if Fade > 0:
        print("You will hit a fade shot with your", entry_name_of_club, round(result_fade, 2), "% of the time out of", len(entry_accuracy.split()), "shots!")
    if Draw > 0:
        print("You will hit a draw shot with your", entry_name_of_club, round(result_draw, 2), "% of the time out of", len(entry_accuracy.split()), "shots!")
    if Push > 0:
        print("You will hit a push shot with your", entry_name_of_club, round(result_push, 2), "% of the time out of", len(entry_accuracy.split()), "shots!")
    if Pull > 0:
        print("You will hit a pull shot with your", entry_name_of_club, round(result_pull, 2), "% of the time out of", len(entry_accuracy.split()), "shots!")

#Pie Chart for the accuracy of a golf club
def accuruacy_pie_chart(accuracy, name_of_club):
    entry_name_of_club = name_of_club.get()
    entry_accuracy = accuracy.get()

    #Counter for the occurrences of the different accuracies in the entry
    Hook = entry_accuracy.count("Hook")
    Slice = entry_accuracy.count("Slice")
    Fade = entry_accuracy.count("Fade")
    Draw = entry_accuracy.count("Draw")
    Push = entry_accuracy.count("Push")
    Pull = entry_accuracy.count("Pull")

    #If statements that will append values to arrays if there is an occurence of one of the accuracies in the users entry
    labels = []
    sizes = []
    if Hook > 0:
        sizes.append(Hook)
        labels.append("Hook")
    if Slice > 0:
        sizes.append(Slice)
        labels.append("Slice")
    if Fade > 0:
        sizes.append(Fade)
        labels.append("Fade")
    if Draw > 0:
        sizes.append(Draw)
        labels.append("Draw")
    if Push > 0:
        sizes.append(Push)
        labels.append("Push")
    if Pull > 0:
        sizes.append(Pull)
        labels.append("Pull")

    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    ax1.axis('equal')
    plt.title('Accuracy of ' + entry_name_of_club)
    plt.show()

#Opens new GUI window to calculate the percentage chance of making a putt at a certain distance
def percentage_chance_of_putt():
    #Main window
    newWindow = Toplevel(master)
    newWindow.title("Calculate Chance of Making a Putt")
    newWindow.geometry("700x200")

    # Header
    main_header = Label(newWindow, text="Chance of Making a Putt")
    main_header.config(font=("Times New Roman", 20))
    main_header.grid(row=0, column=0)

    # Buttons
    Button(newWindow, text="Graph", command=lambda: chance_of_making_putt_pie_chart(length, amount_of_putts)).grid(row=3, column=0)
    Button(newWindow, text="Calculate", command=lambda: calculate_change_of_putt(length, amount_of_putts)).grid(row=3, column=1)

    # Label and entry for percentage chance of making a putt
    Label(newWindow, text='Length from the hole(10ft, 5ft, 1ft, etc)').grid(row=1)
    Label(newWindow, text='Enter the amount of putts it took to make it into the hole (Ex: 1, 2, 3, 4)').grid(row=2)
    length = Entry(newWindow)
    amount_of_putts = Entry(newWindow)
    length.grid(row=1, column=1)
    amount_of_putts.grid(row=2, column=1)

def calculate_change_of_putt(length, amount_of_putts):
    entry_length = length.get()
    entry_amount_of_putts = amount_of_putts.get()

    #Counting the amount of putts inputted
    One_Putt = entry_amount_of_putts.count("1")
    Two_Putt = entry_amount_of_putts.count("2")
    Three_Putt = entry_amount_of_putts.count("3")
    Four_Putt = entry_amount_of_putts.count("4")
    Five_Putt = entry_amount_of_putts.count("5")
    Six_Putt = entry_amount_of_putts.count("6")
    Seven_Putt = entry_amount_of_putts.count("7")
    Eight_Putt = entry_amount_of_putts.count("8")
    Nine_Putt = entry_amount_of_putts.count("9")

    #Calculating the percentages of making a putt
    result_one_putt = (One_Putt / len(entry_amount_of_putts.split()) * 100)
    result_two_putt = (Two_Putt / len(entry_amount_of_putts.split()) * 100)
    result_three_putt = (Three_Putt / len(entry_amount_of_putts.split()) * 100)
    result_four_putt = (Four_Putt / len(entry_amount_of_putts.split()) * 100)
    result_five_putt = (Five_Putt / len(entry_amount_of_putts.split()) * 100)
    result_six_putt = (Six_Putt / len(entry_amount_of_putts.split()) * 100)
    result_seven_putt = (Seven_Putt / len(entry_amount_of_putts.split()) * 100)
    result_eight_putt = (Eight_Putt / len(entry_amount_of_putts.split()) * 100)
    result_nine_putt = (Nine_Putt / len(entry_amount_of_putts.split()) * 100)

    #If statements to check if a person inputed a certain amount of putts, to then print the result of the calculation
    if result_one_putt > 0:
        print("You will have a", round(result_one_putt, 2), "% change of making a one putt from", entry_length, "away!")
    if result_two_putt > 0:
        print("You will have a", round(result_two_putt, 2), "% change of making a two putt from", entry_length, "away!")
    if result_three_putt > 0:
        print("You will have a", round(result_three_putt, 2), "% change of making a three putt from", entry_length, "away!")
    if result_four_putt > 0:
        print("You will have a", round(result_four_putt, 2), "% change of making a four putt from", entry_length, "away!")
    if result_five_putt > 0:
        print("You will have a", round(result_five_putt, 2), "% change of making a five putt from", entry_length, "away!")
    if result_six_putt > 0:
        print("You will have a", round(result_six_putt, 2), "% change of making a six putt from", entry_length, "away!")
    if result_seven_putt > 0:
        print("You will have a", round(result_seven_putt, 2), "% change of making a seven putt from", entry_length, "away!")
    if result_eight_putt > 0:
        print("You will have a", round(result_eight_putt, 2), "% change of making a eight putt from", entry_length, "away!")
    if result_nine_putt > 0:
        print("You will have a", round(result_nine_putt, 2), "% change of making a nine putt from", entry_length, "away!")


# Pie Chart for the percentage chance of making a putt from a certain distance
def chance_of_making_putt_pie_chart(length, amount_of_putts):
    entry_length = length.get()
    entry_amount_of_putts = amount_of_putts.get()

    # Counter for the occurrences of the number of putts to make it into the hole.
    One_Putt = entry_amount_of_putts.count("1")
    Two_Putt = entry_amount_of_putts.count("2")
    Three_Putt = entry_amount_of_putts.count("3")
    Four_Putt = entry_amount_of_putts.count("4")
    Five_Putt = entry_amount_of_putts.count("5")
    Six_Putt = entry_amount_of_putts.count("6")
    Seven_Putt = entry_amount_of_putts.count("7")
    Eight_Putt = entry_amount_of_putts.count("8")
    Nine_Putt = entry_amount_of_putts.count("9")

    #If statements that will append values to arrays if there is an occurence of one of the number of putts it took to make it into the hole
    labels = []
    sizes = []
    if One_Putt > 0:
        sizes.append(One_Putt)
        labels.append("One Putt")
    if Two_Putt > 0:
        sizes.append(Two_Putt)
        labels.append("Two Putt")
    if Three_Putt > 0:
        sizes.append(Three_Putt)
        labels.append("Three Putt")
    if Four_Putt > 0:
        sizes.append(Four_Putt)
        labels.append("Four Putt")
    if Five_Putt > 0:
        sizes.append(Five_Putt)
        labels.append("Five Putt")
    if Six_Putt > 0:
        sizes.append(Six_Putt)
        labels.append("Six Putt")
    if Seven_Putt > 0:
        sizes.append(Seven_Putt)
        labels.append("Seven Putt")
    if Eight_Putt > 0:
        sizes.append(Eight_Putt)
        labels.append("Eight Putt")
    if Nine_Putt > 0:
        sizes.append(Nine_Putt)
        labels.append("Nine Putt")

    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    ax1.axis('equal')
    plt.title("Percentage Chance of Making a Putt from " + entry_length + " Away")
    plt.show()

#Function that calls a file to run a paint program
def tracking_consistency():
    runpy.run_path(path_name='Tracking_Consistency_Paint_File.py')

#Function to read excel file
def read_excel_file():
    df = pd.read_excel('Golf_Statistics.xlsx')
    print(df)

#Work in progress
# df = pd.read_excel ('Golf_Statistics.xlsx')
# def sample_gcode():
#     Label(master, text="Gcode", anchor="w", bg="gray20", fg="lime green", font=('Helvetica', '12')).place(x=1290, y=560, height=20,width=100)
#     Gcode_Text = scrolledtext.ScrolledText(master, bg='white', relief=GROOVE, font='TkFixedFont')
#     Gcode_Text.insert(INSERT, df)
#     Gcode_Text.place(x=1290, y=580, height=120, width=200)

#Function to save data to an excel file and to exit the code
def save_exit():
    sheet1 = wb.worksheets[0]
    sheet1.append(main_array)
    wb.save(filename='Golf_Statistics.xlsx')
    exit()

#Buttons
Button(master, text="Calculate Average Distance of Golf Club",command=average_distance).grid(row=4)
Button(master, text="Calculate the Accuracy of a Golf Club",command=accuracy_of_shot).grid(row=5)
Button(master, text="Calculate the Percentage Chance of Making a Putt",command=percentage_chance_of_putt).grid(row=6)
Button(master, text="Tracking Consistency",command=tracking_consistency).grid(row=7)
Button(master, text="Save and Exit",command=save_exit).grid(row=9, column=1)
Button(master, text="View Xlsx Data",command=read_excel_file).grid(row=9, column=0)
Button(master, text="Enter",command=name_user).grid(row=2, column=2)
Button(master, text="Enter",command=date_of_user).grid(row=3, column=2)
Button(master, text="Enter",command=user_notes).grid(row=8, column=2)

master.mainloop()